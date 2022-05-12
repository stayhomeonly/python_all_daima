"""
=========================
File Name:excel_test05
Author:冯鑫
Date:2021/12/6-12:58
==========================
"""
import openpyxl

'''
将测试数据封装成对象
'''


class CaseData:
    def __init__(self, dict_case):
        for i in dict_case.items():  # 获取字典中的每一对键值对，并且进行遍历
            setattr(self, i[0], i[1])


# 读取所有数据
class ReadExcel:
    @classmethod
    def read_data_all(cls, file_name, sheet_name):  # 读取,文件名，sheet_name中的所有内容
        workbook = openpyxl.load_workbook(file_name)
        sh = workbook[sheet_name]
        rows = list(sh.rows)  # 获取所有单元格
        all_cases = []
        # 读取表头数据
        titles = []
        for cell in rows[0]:  # row[0]代表第一行数据
            titles.append(cell.value)  # 将第一行的cell的值添加到titles中
        # print(titles)  # ['case_id','case_title',interface........]

        # 遍历其他行数据，和表头打包，转换为字典，放到列表

        for row in rows[1:]:
            data = []  # row 代表除了表头的每一行数据
            # [1,'登录接口主流程'，‘login',....]
            for v in row:
                data.append(v.value)
            case_zip = dict(zip(titles, data))  # 将数据打包 {‘case_id’:1,'case_titles':'登录接口主流程'
            cd = CaseData(case_zip)  # 调用反射类，将字典数据转换为对象
            all_cases.append(cd)
        return all_cases

    # 读取部门的内容
    @classmethod
    def read_data_pl(cls, file_name, sheet_name, begin_row, end_row):
        workbook = openpyxl.load_workbook(file_name)  # 读取工作簿
        sh = workbook[sheet_name]  # 读取sheet页的所有内容
        rows = list(sh.rows)  # 获取所有的行数
        all_cases = []

        # 读取表头数据
        titles = []
        for cell in rows[0]:  # row[0]代表第一行数据
            titles.append(cell.value)  # 将第一行的cell的值添加到titles中
        # print(titles)  # ['case_id','case_title',interface........]

        # 遍历其他行数据，和表头打包，转换为字典，放到列表
        for row in rows[begin_row:end_row + 1]:  # row 代表除了表头的每一行数据
            data = []  # [1,'登录接口主流程'，‘login',....]
            for v in row:
                data.append(v.value)
            case_zip = dict(zip(titles, data))  # 将数据打包 {‘case_id’:1,'case_titles':'登录接口主流程'
            cd = CaseData(case_zip)  # 调用反射类，将字典数据转换为对象
            all_cases.append(cd)
        return all_cases

    # 写入的内容
    @classmethod
    def write_data(cls, file_name, sheet_name, row, column, value):
        """

        :param file_name: 需要读写的文件
        :param sheet_name: 需要读写的sheet页
        :param row: 需要被写入的行数
        :param column: 需要被写入的列数
        :param value: 需要被写入的内容

        """

        workbook = openpyxl.load_workbook(file_name)
        sh = workbook[sheet_name]
        sh.cell(row=row + 1, column=column, value=value)
        workbook.save(file_name)
