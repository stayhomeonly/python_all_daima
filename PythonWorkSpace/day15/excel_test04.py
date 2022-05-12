"""
=========================
File Name:cases_xlsx
Author:冯鑫
Date:2021/12/6-11:10
==========================
"""
"""
python操作excel文件
"""

import openpyxl  # 用来操作excel文件

workbook = openpyxl.load_workbook(r'./cases.xlsx')  # 用来读取excel文件并且看作一个对象
sheet = workbook['Login']  # 获取sheet页

# 获取cell中值(excel文件读取的数据类型除了数字类型，其他都是str)
res = sheet.cell(row=2, column=5).value  # cell 单元格的意思，row行，column列
print(res)  # {"username": "xiaowang", "password": "a123456"}
print(type(res))  # str类型
print(list(sheet.rows))  # 获取所有的cell单元格，转换成列表

# 写数据
sheet.cell(row=2,column=7).value = '23班第一次excel读写'
workbook.save(r'./cases.xlsx') # 注意写名字一旦写错就相当于另存为新文件

